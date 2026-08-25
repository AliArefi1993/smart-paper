import datetime
import io

from django.contrib.auth.hashers import make_password
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from finance.models import IncomeEntry

from .models import DayPlan, DayScheduleEntry, PlannerSectionConfig, Week
from .views import saturday_start


class PlannerApiTests(TestCase):
    def test_weeks_list_includes_current_week(self):
        response = self.client.get(reverse("weeks-list"))
        self.assertEqual(response.status_code, 200)
        payload = response.json()

        current = payload["current_week_start"]
        starts = [week["start_date"] for week in payload["weeks"]]
        self.assertIn(current, starts)

    def test_week_detail_creates_week_days(self):
        week_start = "2026-04-25"  # Saturday
        response = self.client.get(reverse("week-detail", args=[week_start]))
        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["start_date"], week_start)
        self.assertEqual(len(payload["days"]), 7)
        self.assertEqual(len(payload["planner_sections"]), 10)
        self.assertEqual(payload["planner_sections"][0]["slot_id"], "slot_1")
        self.assertEqual(payload["planner_sections"][0]["label"], "Main")
        self.assertTrue(payload["planner_sections"][0]["active"])
        self.assertFalse(payload["planner_sections"][-1]["active"])
        self.assertIn("slot_10", payload["days"][0]["sections"])
        self.assertEqual(payload["days"][0]["schedule_entries"], [])
        self.assertEqual(Week.objects.count(), 1)
        self.assertEqual(DayPlan.objects.count(), 7)
        self.assertEqual(PlannerSectionConfig.objects.count(), 10)

    def test_planner_sections_get_put(self):
        response = self.client.get(reverse("planner-sections"))
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload["planner_sections"]), 10)

        response = self.client.put(
            reverse("planner-sections"),
            data={
                "planner_sections": [
                    {"slot_id": "slot_5", "label": "Writing", "active": True},
                    {"slot_id": "slot_2", "label": "Admin", "active": False},
                ]
            },
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        sections_by_slot = {
            section["slot_id"]: section for section in response.json()["planner_sections"]
        }
        self.assertEqual(sections_by_slot["slot_5"]["label"], "Writing")
        self.assertTrue(sections_by_slot["slot_5"]["active"])
        self.assertEqual(sections_by_slot["slot_2"]["label"], "Admin")
        self.assertFalse(sections_by_slot["slot_2"]["active"])

    def test_planner_sections_rejects_deactivating_every_section(self):
        response = self.client.put(
            reverse("planner-sections"),
            data={
                "planner_sections": [
                    {"slot_id": "slot_1", "active": False},
                    {"slot_id": "slot_2", "active": False},
                    {"slot_id": "slot_3", "active": False},
                    {"slot_id": "slot_4", "active": False},
                ]
            },
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertTrue(PlannerSectionConfig.objects.get(slot_id="slot_1").active)

    def test_week_update_happy_path(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))

        update_payload = {
            "weekly_goal": "Finish key tasks for the week",
            "weekly_note": "Keep focus and avoid context switching",
            "days": [
                {
                    "date": "2026-04-25",
                    "sections": {
                        "main": {
                            "duration_minutes": 120,
                            "goal": "Ship backend",
                            "note": "Deep work",
                        },
                        "learning": {
                            "duration_minutes": 45,
                            "goal": "Learn APIs",
                            "note": "Django docs",
                        },
                    },
                },
                {
                    "date": "2026-04-26",
                    "sections": {
                        "exercise": {
                            "duration_minutes": 60,
                            "goal": "Cardio + strength",
                            "note": "Gym",
                        },
                    },
                },
            ]
        }

        response = self.client.put(
            reverse("week-detail", args=[week_start]),
            data=update_payload,
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["weekly_goal"], "Finish key tasks for the week")
        self.assertEqual(
            payload["weekly_note"], "Keep focus and avoid context switching"
        )
        self.assertEqual(payload["days"][0]["sections"]["slot_1"]["goal"], "Ship backend")
        self.assertEqual(payload["totals"]["week_total_minutes"], 225)
        self.assertEqual(payload["totals"]["by_section_minutes"]["slot_1"], 120)
        self.assertEqual(payload["totals"]["by_section_minutes"]["slot_3"], 45)
        self.assertEqual(payload["totals"]["by_section_minutes"]["slot_4"], 60)

    def test_week_update_saves_sorts_and_replaces_schedule_entries(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))

        response = self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "days": [
                    {
                        "date": "2026-04-25",
                        "schedule_entries": [
                            {
                                "start_time": "18:00",
                                "end_time": "19:00",
                                "title": "Meeting",
                                "note": "Discuss launch",
                                "section_id": "slot_1",
                            },
                            {
                                "start_time": "08:30",
                                "end_time": "09:00",
                                "title": "Plan day",
                            },
                        ],
                    }
                ]
            },
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        entries = response.json()["days"][0]["schedule_entries"]
        self.assertEqual([entry["title"] for entry in entries], ["Plan day", "Meeting"])
        self.assertEqual(entries[1]["start_time"], "18:00")
        self.assertEqual(entries[1]["section_id"], "slot_1")
        self.assertEqual(DayScheduleEntry.objects.count(), 2)

        response = self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "days": [
                    {
                        "date": "2026-04-25",
                        "schedule_entries": [
                            {
                                "id": entries[0]["id"],
                                "start_time": "10:00",
                                "end_time": "11:00",
                                "title": "Updated plan",
                                "note": "",
                                "section_id": None,
                            }
                        ],
                    }
                ]
            },
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        entries = response.json()["days"][0]["schedule_entries"]
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0]["title"], "Updated plan")
        self.assertEqual(DayScheduleEntry.objects.count(), 1)

    def test_week_update_rejects_invalid_schedule_entry_time(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))

        response = self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "days": [
                    {
                        "date": "2026-04-25",
                        "schedule_entries": [
                            {
                                "start_time": "19:00",
                                "end_time": "18:00",
                                "title": "Backwards",
                            }
                        ],
                    }
                ]
            },
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(DayScheduleEntry.objects.count(), 0)

    def test_week_update_accepts_slot_ids_and_counts_active_only(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))

        response = self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "days": [
                    {
                        "date": "2026-04-25",
                        "sections": {
                            "slot_1": {"duration_minutes": 15, "goal": "Core"},
                            "slot_5": {
                                "duration_minutes": 90,
                                "goal": "Hidden project",
                                "note": "Keep hidden data",
                            },
                        },
                    }
                ]
            },
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["days"][0]["sections"]["slot_5"]["goal"], "Hidden project")
        self.assertEqual(payload["totals"]["week_total_minutes"], 15)
        self.assertNotIn("slot_5", payload["totals"]["by_section_minutes"])

        self.client.put(
            reverse("planner-sections"),
            data={"planner_sections": [{"slot_id": "slot_5", "active": True}]},
            content_type="application/json",
        )
        response = self.client.get(reverse("week-detail", args=[week_start]))
        self.assertEqual(response.json()["totals"]["week_total_minutes"], 105)
        self.assertEqual(response.json()["totals"]["by_section_minutes"]["slot_5"], 90)

    def test_inactive_section_data_is_not_cleared_by_omitted_save(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))
        self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "days": [
                    {
                        "date": "2026-04-25",
                        "sections": {
                            "slot_5": {
                                "duration_minutes": 50,
                                "goal": "Private backlog",
                                "note": "Preserve me",
                            }
                        },
                    }
                ]
            },
            content_type="application/json",
        )

        response = self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "days": [
                    {
                        "date": "2026-04-25",
                        "sections": {"slot_1": {"duration_minutes": 10}},
                    }
                ]
            },
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        slot_5 = response.json()["days"][0]["sections"]["slot_5"]
        self.assertEqual(slot_5["duration_minutes"], 50)
        self.assertEqual(slot_5["goal"], "Private backlog")
        self.assertEqual(slot_5["note"], "Preserve me")

    def test_week_summaries_happy_path(self):
        week_start = saturday_start(datetime.date.today()).isoformat()
        self.client.get(reverse("week-detail", args=[week_start]))
        first_day = week_start
        update_payload = {
            "weekly_goal": "Stay consistent",
            "days": [
                {
                    "date": first_day,
                    "sections": {
                        "main": {"duration_minutes": 30, "note": "Kickoff"},
                        "exercise": {"duration_minutes": 20, "note": "Stretch"},
                    },
                }
            ],
        }
        self.client.put(
            reverse("week-detail", args=[week_start]),
            data=update_payload,
            content_type="application/json",
        )

        response = self.client.get(reverse("week-summaries") + "?span=1")
        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertIn("summaries", payload)
        self.assertGreaterEqual(len(payload["summaries"]), 3)

        target = next(
            item for item in payload["summaries"] if item["start_date"] == week_start
        )
        self.assertEqual(target["weekly_goal"], "Stay consistent")
        self.assertEqual(target["totals"]["week_total_minutes"], 50)
        self.assertIn("planner_sections", target)
        self.assertIn("Saturday: Kickoff", target["notes_by_section"]["slot_1"])
        self.assertEqual(target["details_by_section"]["slot_1"][0]["goal"], "")
        self.assertEqual(target["details_by_section"]["slot_1"][0]["note"], "Kickoff")

    def test_export_all_data_requires_finance_unlock(self):
        response = self.client.get(reverse("export-all-data"))
        self.assertEqual(response.status_code, 403)

    def test_export_all_data_json_happy_path(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))
        self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "weekly_goal": "Exportable week",
                "days": [
                    {
                        "date": "2026-04-25",
                        "sections": {
                            "main": {
                                "duration_minutes": 90,
                                "goal": "Make export",
                                "note": "Backup ready",
                            },
                        },
                    }
                ],
            },
            content_type="application/json",
        )

        with self.settings(FINANCE_PIN_HASH=make_password("1234")):
            self.client.post(
                reverse("finance-unlock"),
                data={"pin": "1234"},
                content_type="application/json",
            )
            self.client.put(
                reverse("finance-overview"),
                data={
                    "goal_amount": 10000,
                    "income_amount": 1200,
                    "income_note": "Project",
                    "income_date": "2026-04-26",
                },
                content_type="application/json",
            )

            response = self.client.get(reverse("export-all-data"))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["schema_version"], 3)
        self.assertIn("planner_sections", payload)
        self.assertIn("exported_at", payload)
        self.assertEqual(payload["weeks"][0]["weekly_goal"], "Exportable week")
        self.assertEqual(payload["weeks"][0]["totals"]["week_total_minutes"], 90)
        self.assertEqual(payload["finance"]["goal_amount"], 10000)
        self.assertEqual(payload["finance"]["entries"][0]["note"], "Project")

    def test_export_all_data_csv_happy_path(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))
        self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "weekly_goal": "CSV week",
                "days": [
                    {
                        "date": "2026-04-25",
                        "sections": {
                            "main": {
                                "duration_minutes": 30,
                                "goal": "CSV export",
                                "note": "Sheet row",
                            },
                        },
                    }
                ],
            },
            content_type="application/json",
        )

        with self.settings(FINANCE_PIN_HASH=make_password("1234")):
            self.client.post(
                reverse("finance-unlock"),
                data={"pin": "1234"},
                content_type="application/json",
            )
            self.client.put(
                reverse("finance-overview"),
                data={
                    "income_amount": 700,
                    "income_note": "CSV income",
                    "income_date": "2026-04-27",
                },
                content_type="application/json",
            )

            response = self.client.get(reverse("export-all-data") + "?format=csv")

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("record_type,week_start", content)
        self.assertIn("day_section,2026-04-25", content)
        self.assertIn("CSV export", content)
        self.assertIn("income,", content)
        self.assertIn("CSV income", content)

    def test_export_all_data_includes_schedule_entries(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))
        self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "days": [
                    {
                        "date": "2026-04-25",
                        "schedule_entries": [
                            {
                                "start_time": "18:00",
                                "end_time": "19:00",
                                "title": "Release meeting",
                                "note": "Bring notes",
                                "section_id": "slot_1",
                            }
                        ],
                    }
                ],
            },
            content_type="application/json",
        )

        with self.settings(FINANCE_PIN_HASH=make_password("1234")):
            self.client.post(
                reverse("finance-unlock"),
                data={"pin": "1234"},
                content_type="application/json",
            )
            response = self.client.get(reverse("export-all-data"))

        self.assertEqual(response.status_code, 200)
        entry = response.json()["weeks"][0]["days"][0]["schedule_entries"][0]
        self.assertEqual(entry["title"], "Release meeting")
        self.assertEqual(entry["start_time"], "18:00")

    def test_export_all_data_xlsx_happy_path(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))
        self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "weekly_goal": "Readable Excel week",
                "days": [
                    {
                        "date": "2026-04-25",
                        "sections": {
                            "main": {
                                "duration_minutes": 55,
                                "goal": "Excel export",
                                "note": "Readable row",
                            },
                        },
                    }
                ],
            },
            content_type="application/json",
        )

        with self.settings(FINANCE_PIN_HASH=make_password("1234")):
            self.client.post(
                reverse("finance-unlock"),
                data={"pin": "1234"},
                content_type="application/json",
            )
            self.client.put(
                reverse("finance-overview"),
                data={
                    "income_amount": 900,
                    "income_note": "Excel income",
                    "income_date": "2026-04-28",
                },
                content_type="application/json",
            )

            response = self.client.get(reverse("export-all-data") + "?format=xlsx")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames, ["Overview", "Weeks", "Day Sections", "Schedule", "Income"]
        )
        self.assertEqual(workbook["Weeks"]["C2"].value, "Readable Excel week")
        self.assertEqual(workbook["Day Sections"]["E2"].value, 55)
        self.assertEqual(workbook["Income"]["C2"].value, "Excel income")

    def test_export_all_data_markdown_happy_path(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))
        self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "weekly_goal": "Get AI feedback",
                "days": [
                    {
                        "date": "2026-04-25",
                        "sections": {
                            "main": {
                                "duration_minutes": 40,
                                "goal": "Write context",
                                "note": "Ask better questions",
                            },
                        },
                    }
                ],
            },
            content_type="application/json",
        )

        with self.settings(FINANCE_PIN_HASH=make_password("1234")):
            self.client.post(
                reverse("finance-unlock"),
                data={"pin": "1234"},
                content_type="application/json",
            )
            response = self.client.get(
                reverse("export-all-data") + "?format=markdown"
            )

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("# Smart Paper Export", content)
        self.assertIn("## AI Feedback Request", content)
        self.assertIn("Get AI feedback", content)
        self.assertIn("Ask better questions", content)

    def test_import_all_data_merge_upserts_without_deleting_existing_data(self):
        existing_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[existing_start]))
        self.client.put(
            reverse("week-detail", args=[existing_start]),
            data={
                "weekly_goal": "Old goal",
                "days": [
                    {
                        "date": existing_start,
                        "sections": {
                            "main": {
                                "duration_minutes": 10,
                                "goal": "Old work",
                                "note": "Keep me updated",
                            },
                        },
                    }
                ],
            },
            content_type="application/json",
        )

        with self.settings(FINANCE_PIN_HASH=make_password("1234")):
            self.client.post(
                reverse("finance-unlock"),
                data={"pin": "1234"},
                content_type="application/json",
            )
            self.client.put(
                reverse("finance-overview"),
                data={
                    "income_amount": 300,
                    "income_note": "Existing income",
                    "income_date": "2026-04-25",
                },
                content_type="application/json",
            )
            existing_income_id = IncomeEntry.objects.get().id

            response = self.client.post(
                reverse("import-all-data") + "?mode=merge",
                data={
                    "exported_at": "2026-04-30T00:00:00+00:00",
                    "schema_version": 1,
                    "weeks": [
                        {
                            "start_date": existing_start,
                            "weekly_goal": "Imported goal",
                            "weekly_note": "Merged note",
                            "days": [
                                {
                                    "date": existing_start,
                                    "sections": {
                                        "main": {
                                            "duration_minutes": 45,
                                            "goal": "Imported work",
                                            "note": "Updated by import",
                                        }
                                    },
                                }
                            ],
                        }
                    ],
                    "finance": {
                        "goal_amount": 5000,
                        "entries": [
                            {
                                "id": existing_income_id,
                                "amount": 450,
                                "note": "Updated income",
                                "received_on": "2026-04-26",
                            },
                            {
                                "id": 98765,
                                "amount": 900,
                                "note": "New import income",
                                "received_on": "2026-04-27",
                            },
                        ],
                    },
                },
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["weeks_imported"], 1)
        week = Week.objects.get(start_date=existing_start)
        self.assertEqual(week.weekly_goal, "Imported goal")
        saturday = week.days.get(date=existing_start)
        self.assertEqual(saturday.main_duration_minutes, 45)
        self.assertEqual(IncomeEntry.objects.count(), 2)
        self.assertEqual(IncomeEntry.objects.get(id=existing_income_id).amount, 450)
        self.assertTrue(IncomeEntry.objects.filter(id=98765).exists())

    def test_import_all_data_replaces_schedule_entries_for_imported_day(self):
        week_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[week_start]))
        self.client.put(
            reverse("week-detail", args=[week_start]),
            data={
                "days": [
                    {
                        "date": week_start,
                        "schedule_entries": [
                            {
                                "start_time": "08:00",
                                "end_time": "09:00",
                                "title": "Old schedule",
                            }
                        ],
                    }
                ],
            },
            content_type="application/json",
        )

        with self.settings(FINANCE_PIN_HASH=make_password("1234")):
            self.client.post(
                reverse("finance-unlock"),
                data={"pin": "1234"},
                content_type="application/json",
            )
            response = self.client.post(
                reverse("import-all-data") + "?mode=merge",
                data={
                    "exported_at": "2026-04-30T00:00:00+00:00",
                    "schema_version": 3,
                    "weeks": [
                        {
                            "start_date": week_start,
                            "weekly_goal": "",
                            "weekly_note": "",
                            "days": [
                                {
                                    "date": week_start,
                                    "sections": {},
                                    "schedule_entries": [
                                        {
                                            "start_time": "18:00",
                                            "end_time": "19:00",
                                            "title": "Imported meeting",
                                            "note": "New schedule",
                                            "section_id": "slot_2",
                                        }
                                    ],
                                }
                            ],
                        }
                    ],
                    "finance": {"goal_amount": 0, "entries": []},
                },
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        day = Week.objects.get(start_date=week_start).days.get(date=week_start)
        entries = list(day.schedule_entries.all())
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0].title, "Imported meeting")
        self.assertEqual(entries[0].start_minutes, 18 * 60)

    def test_import_all_data_replace_deletes_old_data_first(self):
        old_start = "2026-04-25"
        self.client.get(reverse("week-detail", args=[old_start]))

        with self.settings(FINANCE_PIN_HASH=make_password("1234")):
            self.client.post(
                reverse("finance-unlock"),
                data={"pin": "1234"},
                content_type="application/json",
            )
            self.client.put(
                reverse("finance-overview"),
                data={
                    "income_amount": 300,
                    "income_note": "Old income",
                    "income_date": "2026-04-25",
                },
                content_type="application/json",
            )

            response = self.client.post(
                reverse("import-all-data") + "?mode=replace",
                data={
                    "exported_at": "2026-04-30T00:00:00+00:00",
                    "schema_version": 2,
                    "planner_sections": [
                        {"slot_id": "slot_5", "label": "Writing", "active": True}
                    ],
                    "weeks": [
                        {
                            "start_date": "2026-05-02",
                            "weekly_goal": "Only imported week",
                            "weekly_note": "",
                            "days": [
                                {
                                    "date": "2026-05-02",
                                    "sections": {
                                        "slot_5": {
                                            "duration_minutes": 25,
                                            "goal": "Imported slot",
                                            "note": "Visible after config",
                                        }
                                    },
                                }
                            ],
                        }
                    ],
                    "finance": {
                        "goal_amount": 1000,
                        "entries": [
                            {
                                "id": 222,
                                "amount": 100,
                                "note": "Only imported income",
                                "received_on": "2026-05-02",
                            }
                        ],
                    },
                },
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Week.objects.filter(start_date=old_start).exists())
        self.assertTrue(Week.objects.filter(start_date="2026-05-02").exists())
        section = PlannerSectionConfig.objects.get(slot_id="slot_5")
        self.assertEqual(section.label, "Writing")
        self.assertTrue(section.active)
        imported_week = Week.objects.get(start_date="2026-05-02")
        self.assertEqual(
            imported_week.days.get(date="2026-05-02").slot_5_duration_minutes,
            25,
        )
        self.assertEqual(IncomeEntry.objects.count(), 1)
        self.assertEqual(IncomeEntry.objects.get().note, "Only imported income")

    def test_import_all_data_accepts_local_mode_section_shape(self):
        with self.settings(FINANCE_PIN_HASH=make_password("1234")):
            self.client.post(
                reverse("finance-unlock"),
                data={"pin": "1234"},
                content_type="application/json",
            )
            response = self.client.post(
                reverse("import-all-data") + "?mode=merge",
                data={
                    "exported_at": "2026-04-30T00:00:00+00:00",
                    "schema_version": 2,
                    "planner_sections": [
                        {"id": "slot_5", "label": "Writing", "active": True, "order": 5}
                    ],
                    "weeks": [],
                    "finance": {"goal_amount": 0, "entries": []},
                },
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        section = PlannerSectionConfig.objects.get(slot_id="slot_5")
        self.assertEqual(section.label, "Writing")
        self.assertTrue(section.active)

    def test_import_all_data_keeps_at_least_one_planner_section_active(self):
        with self.settings(FINANCE_PIN_HASH=make_password("1234")):
            self.client.post(
                reverse("finance-unlock"),
                data={"pin": "1234"},
                content_type="application/json",
            )
            response = self.client.post(
                reverse("import-all-data") + "?mode=merge",
                data={
                    "exported_at": "2026-04-30T00:00:00+00:00",
                    "schema_version": 2,
                    "planner_sections": [
                        {"slot_id": "slot_1", "active": False},
                        {"slot_id": "slot_2", "active": False},
                        {"slot_id": "slot_3", "active": False},
                        {"slot_id": "slot_4", "active": False},
                    ],
                    "weeks": [],
                    "finance": {"goal_amount": 0, "entries": []},
                },
                content_type="application/json",
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(PlannerSectionConfig.objects.get(slot_id="slot_1").active)
