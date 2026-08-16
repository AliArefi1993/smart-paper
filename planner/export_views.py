import csv
import datetime
import io
import json

from django.db import transaction
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from finance.models import FinanceState, IncomeEntry
from finance.views import require_finance_unlock, serialize_finance

from .models import DayPlan, PlannerSectionConfig, Week
from .views import (
    SECTION_FIELD_NAMES,
    SECTIONS,
    SLOT_IDS,
    WEEKDAY_NAMES,
    build_week,
    field_prefix_for_slot,
    normalize_section_key,
    serialize_planner_sections,
    serialize_week,
)

SCHEMA_VERSION = 2


EXPORT_COLUMNS = [
    "record_type",
    "week_start",
    "week_end",
    "date",
    "weekday",
    "section",
    "section_label",
    "duration_minutes",
    "goal",
    "note",
    "finance_goal_amount",
    "income_id",
    "income_amount",
    "income_note",
    "income_date",
]


def export_payload() -> dict:
    planner_sections = serialize_planner_sections()
    weeks = [
        serialize_week(week)
        for week in Week.objects.all().order_by("start_date").prefetch_related("days")
    ]
    return {
        "schema_version": SCHEMA_VERSION,
        "exported_at": datetime.datetime.now(datetime.UTC).isoformat(),
        "planner_sections": planner_sections,
        "weeks": weeks,
        "finance": serialize_finance(),
    }


def section_label_lookup(payload: dict) -> dict[str, str]:
    return {
        section["slot_id"]: section["label"]
        for section in payload.get("planner_sections", [])
        if isinstance(section, dict)
        and section.get("slot_id") in SLOT_IDS
        and isinstance(section.get("label"), str)
    }


def readable_section_ids(payload: dict) -> tuple[str, ...]:
    active_sections = [
        section.get("slot_id")
        for section in payload.get("planner_sections", [])
        if isinstance(section, dict)
        and section.get("slot_id") in SLOT_IDS
        and section.get("active")
    ]
    return tuple(active_sections) or SECTIONS[:4]


def export_csv(payload: dict) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()
    section_labels = section_label_lookup(payload)
    section_ids = readable_section_ids(payload)

    for week in payload["weeks"]:
        writer.writerow(
            {
                "record_type": "week",
                "week_start": week["start_date"],
                "week_end": week["end_date"],
                "goal": week["weekly_goal"],
                "note": week["weekly_note"],
            }
        )
        for day in week["days"]:
            for section in section_ids:
                section_data = day["sections"].get(section)
                if section_data is None:
                    continue
                writer.writerow(
                    {
                        "record_type": "day_section",
                        "week_start": week["start_date"],
                        "week_end": week["end_date"],
                        "date": day["date"],
                        "weekday": WEEKDAY_NAMES[day["weekday_index"]],
                        "section": section,
                        "section_label": section_labels.get(section, section),
                        "duration_minutes": section_data["duration_minutes"],
                        "goal": section_data["goal"],
                        "note": section_data["note"],
                    }
                )

    finance = payload["finance"]
    writer.writerow(
        {
            "record_type": "finance_summary",
            "finance_goal_amount": finance["goal_amount"],
            "income_amount": finance["total_income"],
            "note": f"Remaining: {finance['remaining_amount']}",
        }
    )
    for entry in finance["entries"]:
        writer.writerow(
            {
                "record_type": "income",
                "income_id": entry["id"],
                "income_amount": entry["amount"],
                "income_note": entry["note"],
                "income_date": entry["received_on"],
            }
        )

    return output.getvalue()


def export_markdown(payload: dict) -> str:
    section_labels = section_label_lookup(payload)
    section_ids = readable_section_ids(payload)
    lines = [
        "# Smart Paper Export",
        "",
        f"Exported at: {payload['exported_at']}",
        "",
        "## AI Feedback Request",
        "",
        "Please review my planner and finance data. Look for patterns, risks, missing follow-through, and practical next actions. Ask clarifying questions if needed.",
        "",
        "## Finance",
        "",
        f"- Goal amount: {payload['finance']['goal_amount']}",
        f"- Total income: {payload['finance']['total_income']}",
        f"- Remaining amount: {payload['finance']['remaining_amount']}",
        f"- Progress percent: {payload['finance']['progress_percent']}",
        "",
        "### Income Entries",
        "",
    ]

    if payload["finance"]["entries"]:
        for entry in payload["finance"]["entries"]:
            note = entry["note"] or "No note"
            lines.append(
                f"- {entry['received_on']}: {entry['amount']} ({note})"
            )
    else:
        lines.append("- No income entries.")

    lines.extend(["", "## Planner Weeks", ""])
    for week in payload["weeks"]:
        lines.extend(
            [
                f"### {week['start_date']} to {week['end_date']}",
                "",
                f"- Weekly goal: {week['weekly_goal'] or 'No weekly goal'}",
                f"- Weekly note: {week['weekly_note'] or 'No weekly note'}",
                f"- Total minutes: {week['totals']['week_total_minutes']}",
                "",
            ]
        )

        for day in week["days"]:
            day_lines = []
            for section in section_ids:
                section_data = day["sections"].get(section)
                if section_data is None:
                    continue
                if (
                    section_data["duration_minutes"] == 0
                    and not section_data["goal"]
                    and not section_data["note"]
                ):
                    continue
                section_label = section_labels.get(section, section)
                day_lines.append(
                    f"  - {section} ({section_label}): {section_data['duration_minutes']} min; goal: {section_data['goal'] or 'none'}; note: {section_data['note'] or 'none'}"
                )
            if not day_lines:
                continue
            lines.append(f"- {day['weekday_name']} {day['date']}")
            lines.extend(day_lines)
        lines.append("")

    return "\n".join(lines).strip() + "\n"


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(width + 2, 12), 45
        )


def export_xlsx(payload: dict) -> bytes:
    section_labels = section_label_lookup(payload)
    section_ids = readable_section_ids(payload)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Metric", "Value"])
    overview.append(["Exported at", payload["exported_at"]])
    overview.append(["Weeks", len(payload["weeks"])])
    overview.append(["Finance goal", payload["finance"]["goal_amount"]])
    overview.append(["Total income", payload["finance"]["total_income"]])
    overview.append(["Remaining", payload["finance"]["remaining_amount"]])
    overview.append(["Progress percent", payload["finance"]["progress_percent"]])

    weeks_sheet = workbook.create_sheet("Weeks")
    section_columns = [
        f"{section_labels.get(slot_id, slot_id)} ({slot_id}) minutes"
        for slot_id in section_ids
    ]
    weeks_sheet.append(
        [
            "Week start",
            "Week end",
            "Weekly goal",
            "Weekly note",
            *section_columns,
            "Total minutes",
        ]
    )
    day_sheet = workbook.create_sheet("Day Sections")
    day_sheet.append(
        [
            "Week start",
            "Date",
            "Weekday",
            "Section",
            "Duration minutes",
            "Goal",
            "Note",
        ]
    )

    for week in payload["weeks"]:
        totals = week["totals"]["by_section_minutes"]
        weeks_sheet.append(
            [
                week["start_date"],
                week["end_date"],
                week["weekly_goal"],
                week["weekly_note"],
                *[totals.get(slot_id, 0) for slot_id in section_ids],
                week["totals"]["week_total_minutes"],
            ]
        )
        for day in week["days"]:
            for section in section_ids:
                section_data = day["sections"].get(section)
                if section_data is None:
                    continue
                day_sheet.append(
                    [
                        week["start_date"],
                        day["date"],
                        day["weekday_name"],
                        f"{section} ({section_labels.get(section, section)})",
                        section_data["duration_minutes"],
                        section_data["goal"],
                        section_data["note"],
                    ]
                )

    income_sheet = workbook.create_sheet("Income")
    income_sheet.append(["Date", "Amount", "Note", "Entry ID"])
    for entry in payload["finance"]["entries"]:
        income_sheet.append(
            [entry["received_on"], entry["amount"], entry["note"], entry["id"]]
        )

    for sheet in workbook.worksheets:
        style_sheet(sheet)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_iso_date(value, field_name: str) -> datetime.date:
    if not isinstance(value, str):
        raise ValueError(f"{field_name} must be YYYY-MM-DD")
    try:
        return datetime.date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(f"{field_name} must be YYYY-MM-DD") from exc


def import_week(week_data: dict) -> bool:
    week_start = parse_iso_date(week_data.get("start_date"), "week start_date")
    week = build_week(week_start)

    weekly_goal = week_data.get("weekly_goal")
    weekly_note = week_data.get("weekly_note")
    if isinstance(weekly_goal, str):
        week.weekly_goal = weekly_goal.strip()
    if isinstance(weekly_note, str):
        week.weekly_note = weekly_note.strip()
    week.save(update_fields=["weekly_goal", "weekly_note", "updated_at"])

    days_by_date = {day.date.isoformat(): day for day in week.days.all()}
    changed_days = []
    for day_data in week_data.get("days", []):
        if not isinstance(day_data, dict):
            continue
        day = days_by_date.get(day_data.get("date"))
        if day is None:
            continue
        sections = day_data.get("sections", {})
        if not isinstance(sections, dict):
            continue

        for raw_section, section_data in sections.items():
            section = normalize_section_key(raw_section)
            if section is None:
                continue
            if not isinstance(section_data, dict):
                continue
            field_prefix = field_prefix_for_slot(section)
            duration = section_data.get("duration_minutes")
            goal = section_data.get("goal")
            note = section_data.get("note")
            if isinstance(duration, int) and duration >= 0:
                setattr(day, f"{field_prefix}_duration_minutes", duration)
            if isinstance(goal, str):
                setattr(day, f"{field_prefix}_goal", goal.strip())
            if isinstance(note, str):
                setattr(day, f"{field_prefix}_note", note.strip())
        changed_days.append(day)

    if changed_days:
        DayPlan.objects.bulk_update(changed_days, SECTION_FIELD_NAMES)

    return True


def import_planner_sections(sections_data) -> int:
    if not isinstance(sections_data, list):
        return 0

    existing_by_slot = {
        section.slot_id: section
        for section in PlannerSectionConfig.objects.filter(slot_id__in=SLOT_IDS)
    }
    changed = []
    now = timezone.now()
    for item in sections_data:
        if not isinstance(item, dict):
            continue
        slot_id = item.get("slot_id") or item.get("id")
        if slot_id not in SLOT_IDS or slot_id not in existing_by_slot:
            continue
        section = existing_by_slot[slot_id]
        did_change = False
        label = item.get("label")
        if isinstance(label, str):
            section.label = label.strip() or section.label
            did_change = True
        active = item.get("active")
        if isinstance(active, bool):
            section.active = active
            did_change = True
        if did_change:
            section.updated_at = now
            changed.append(section)

    if changed:
        next_active_by_slot = {
            slot_id: section.active for slot_id, section in existing_by_slot.items()
        }
        for section in changed:
            next_active_by_slot[section.slot_id] = section.active
        if not any(next_active_by_slot.values()):
            first_section = existing_by_slot[SLOT_IDS[0]]
            first_section.active = True
            first_section.updated_at = now
            if first_section not in changed:
                changed.append(first_section)
        PlannerSectionConfig.objects.bulk_update(changed, ["label", "active", "updated_at"])
    return len(changed)


def import_finance(finance_data: dict) -> tuple[bool, int]:
    state, _ = FinanceState.objects.get_or_create(id=1)
    goal_amount = finance_data.get("goal_amount")
    if isinstance(goal_amount, int) and goal_amount >= 0:
        state.goal_amount = goal_amount
        state.save(update_fields=["goal_amount", "updated_at"])

    imported_entries = 0
    for entry_data in finance_data.get("entries", []):
        if not isinstance(entry_data, dict):
            continue
        amount = entry_data.get("amount")
        if not isinstance(amount, int) or amount <= 0:
            continue
        received_on = parse_iso_date(entry_data.get("received_on"), "income received_on")
        note = entry_data.get("note", "")
        if not isinstance(note, str):
            note = ""

        entry_id = entry_data.get("id")
        defaults = {
            "amount": amount,
            "note": note.strip(),
            "received_on": received_on,
        }
        if isinstance(entry_id, int) and entry_id > 0:
            IncomeEntry.objects.update_or_create(id=entry_id, defaults=defaults)
        else:
            IncomeEntry.objects.create(**defaults)
        imported_entries += 1

    return isinstance(goal_amount, int), imported_entries


def import_payload(payload: dict, mode: str) -> dict:
    if mode not in {"merge", "replace"}:
        raise ValueError("mode must be merge or replace")
    if not isinstance(payload.get("weeks"), list) or not isinstance(
        payload.get("finance"), dict
    ):
        raise ValueError("Import file must be a Smart Paper JSON backup")

    with transaction.atomic():
        if mode == "replace":
            DayPlan.objects.all().delete()
            Week.objects.all().delete()
            PlannerSectionConfig.objects.all().delete()
            IncomeEntry.objects.all().delete()
            FinanceState.objects.all().delete()

        serialize_planner_sections()
        planner_sections_imported = import_planner_sections(
            payload.get("planner_sections")
        )

        imported_weeks = 0
        for week_data in payload["weeks"]:
            if not isinstance(week_data, dict):
                continue
            if import_week(week_data):
                imported_weeks += 1

        updated_goal, imported_entries = import_finance(payload["finance"])

    return {
        "mode": mode,
        "weeks_imported": imported_weeks,
        "planner_sections_imported": planner_sections_imported,
        "income_entries_imported": imported_entries,
        "finance_goal_updated": updated_goal,
        "payload": export_payload(),
    }


@require_http_methods(["GET"])
def export_all_data(request):
    locked_response = require_finance_unlock(request)
    if locked_response is not None:
        return locked_response

    export_format = request.GET.get("format", "json")
    payload = export_payload()

    if export_format == "json":
        response = JsonResponse(payload)
        response["Content-Disposition"] = 'attachment; filename="smart-paper-export.json"'
        return response

    if export_format == "csv":
        response = HttpResponse(export_csv(payload), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="smart-paper-export.csv"'
        return response

    if export_format == "xlsx":
        response = HttpResponse(
            export_xlsx(payload),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="smart-paper-export.xlsx"'
        return response

    if export_format in {"markdown", "md"}:
        response = HttpResponse(export_markdown(payload), content_type="text/markdown")
        response["Content-Disposition"] = 'attachment; filename="smart-paper-ai-review.md"'
        return response

    return HttpResponseBadRequest("format must be json, csv, xlsx, or markdown")


@csrf_exempt
@require_http_methods(["POST"])
def import_all_data(request):
    locked_response = require_finance_unlock(request)
    if locked_response is not None:
        return locked_response

    mode = request.GET.get("mode", "merge")
    try:
        payload = json.loads(request.body or "{}")
        result = import_payload(payload, mode)
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON payload")
    except ValueError as exc:
        return HttpResponseBadRequest(str(exc))

    return JsonResponse(result)
